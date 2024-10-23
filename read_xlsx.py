from openpyxl import load_workbook


def read_xlsx(file_name):
    # Excel dosyasını aç
    workbook = load_workbook(file_name)

    # İlk sayfayı al
    sheet = workbook.active

    # Verileri saklamak için bir liste
    data = []

    # Tüm satırları oku
    for row in sheet.iter_rows(values_only=True):
        # Eğer satırda en az bir değer varsa, boş satırları filtrele
        if any(cell is not None for cell in row):
            # Satırı tuple yerine listeye çevir
            data.append(list(row))

    return data


# # Test etme kısmı
# if __name__ == "__main__":
#     # 'Kullanicilar.xlsx' dosyasını kullanarak verileri oku
#     # Excel dosyanızın adı
#     datas = read_xlsx("Kullanicilar.xlsx")
#     for data in datas:
#         print(f"kullanci adi : {
#               data[0]} - sifre : {data[1]} - yetki : {data[2]}")
