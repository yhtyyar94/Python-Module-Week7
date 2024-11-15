from openpyxl import load_workbook


def read_xlsx(file_name):
    # Load the .xlsx file
    workbook = load_workbook(filename=file_name)

    # Select the active worksheet
    sheet = workbook.active

    rows = []

    # Loop through all rows and columns and filter empty rows
    for row in sheet.iter_rows(min_row=1, values_only=True):
        # Check if the row is completely empty
        if all(cell is None for cell in row):
            continue
        rows.append(list(row))  # Convert tuple to list

    if not rows:
        return []

    # Extract headers and filter out None values
    headers = rows[0]
    filtered_headers = [header for header in headers if header is not None]

    # Adjust the rest of the rows based on the filtered headers
    adjusted_rows = [filtered_headers]
    for row in rows[1:]:
        adjusted_row = []
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i] is not None:
                adjusted_row.append(cell)
            elif i < len(headers) and headers[i] is None:
                continue
            else:
                adjusted_row.append(None)
        adjusted_rows.append(adjusted_row)

    return adjusted_rows


if __name__ == "__main__":
    print(read_xlsx("Mentor.xlsx"))
