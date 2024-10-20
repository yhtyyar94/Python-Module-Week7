from openpyxl import load_workbook


def write_xlsx(file_name, rows):
    # Load the .xlsx file
    workbook = load_workbook(filename=file_name)

    # Select the active worksheet
    sheet = workbook.active

    # Loop through all rows and columns and filter empty rows
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            sheet.cell(row=i + 1, column=j + 1, value=cell)

    workbook.save(file_name)
